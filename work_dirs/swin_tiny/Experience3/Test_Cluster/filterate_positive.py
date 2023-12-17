from openpyxl import load_workbook
from collections import Counter
import matplotlib.pyplot as plt
import os
import shutil
# 打开 Excel 文件
workbook = load_workbook('/data4/lj/Classification/1.7/Negative_folder.xlsx')
# workbook = load_workbook('/data4/lj/Classification/1.7/redundanceBox.xlsx')
# 选择要读取的工作表
sheet = workbook['Sheet']  # 替换为你的工作表名称

# 选择要读取的列
columnA = sheet['A']  # 替换为你的列标识，比如'A'代表第一列
columnB = sheet['B']  # 替换为你的列标识，比如'A'代表第一列
# columnE = sheet['E']  # 替换为你的列标识，比如'A'代表第一列
# columnG = sheet['G']  # 替换为你的列标识，比如'A'代表第一列
# column = column_1 - column_2
# 保存列数据的列表
data_list = []

# path = "/data4/lj/Dataset/JML/Experience3/redundanceBox"
path = "/data4/lj/Dataset/JML/Experience3/Negative_folder"
destination_path = "/data4/lj/Dataset/JML/Experience3/TrainSet/negative"
# # 遍历列中的单元格，并将数据添加到列表中
# for i,cell in enumerate(column_1):
#     data_list.append(round(cell.value-column_2[i].value,2))
    
# 遍历列中的单元格，并将数据添加到列表中
for i,cell in enumerate(columnA):
    conf = columnB[i].value
    if conf < 0.5:
        img_path = os.path.join(path,cell.value)
        output_img_path = os.path.join(destination_path,"negative_"+os.path.basename(cell.value))
        shutil.copy(img_path, output_img_path)
    # data_list.append(round(cell.value,2))
# data_list = sorted(data_list)
# Plot1 = Counter(data_list)
# values1 = list(Plot1.keys())
# counts1 = list(Plot1.values())

# 打印列表中的数据
