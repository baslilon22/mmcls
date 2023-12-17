from openpyxl import load_workbook
from collections import Counter
import matplotlib.pyplot as plt
# 打开 Excel 文件
workbook = load_workbook('/data4/lj/Classification/1.7/Negative_folder.xlsx')
# workbook = load_workbook('/data4/lj/Classification/1.7/redundanceBox.xlsx')
# 选择要读取的工作表
sheet = workbook['Sheet']  # 替换为你的工作表名称

# 选择要读取的列
columnA = sheet['A']  # 替换为你的列标识，比如'A'代表第一列
columnB = sheet['B']  # 替换为你的列标识，比如'A'代表第一列
columnE = sheet['E']  # 替换为你的列标识，比如'A'代表第一列
# columnG = sheet['G']  # 替换为你的列标识，比如'A'代表第一列
# column = column_1 - column_2
# 保存列数据的列表
data_list = []


# 遍历列中的单元格，并将数据添加到列表中
for i,cell in enumerate(columnB):
    data_list.append(round(cell.value-columnE[i].value,2))
    
# # 遍历列中的单元格，并将数据添加到列表中
# for cell in columnB:
#     data_list.append(round(cell.value,2))
data_list = sorted(data_list)
Plot1 = Counter(data_list)
values1 = list(Plot1.keys())
counts1 = list(Plot1.values())

#test
test =[]
for i in range(len(data_list)):
    if data_list[i] > 0.3:
        test.append(data_list[i])
        # 16023 19096
        
        
plt.plot(values1,counts1, marker='o',label='Plot1')
plt.legend()
# 添加标题和坐标轴标签
plt.title('Element Frequency')
plt.xlabel('Value')
plt.ylabel('Frequency')
plt.savefig('Threshold_Negative.png')  # 指定文件名和格式，例如PNG
# plt.savefig('D_Value_Negative_4.png')  # 指定文件名和格式，例如PNG

# 打印列表中的数据
print(data_list)