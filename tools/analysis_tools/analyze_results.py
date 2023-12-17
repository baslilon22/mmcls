# Copyright (c) OpenMMLab. All rights reserved.
import argparse
import os.path as osp


import mmcv
from mmcv import DictAction
import openpyxl
from mmcls.datasets import build_dataset
from mmcls.models import build_classifier
import pandas as pd
from sklearn.metrics import confusion_matrix
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import numpy as np
import cv2
from openpyxl.styles import Alignment

dataset_root = "/home/common/linjie/Dataset/Cls_Realist_Summary_Crop_1/train"

def parse_args():
    parser = argparse.ArgumentParser(
        description='MMCls evaluate prediction success/fail')
    parser.add_argument('config', help='test config file path')
    parser.add_argument('result', help='test result json/pkl file')
    parser.add_argument('--out-dir', help='dir to store output files')
    parser.add_argument(
        '--topk',
        default=50,
        type=int,
        help='Number of images to select for success/fail')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
    args = parser.parse_args()

    return args


def Insert_image(file_name,row_index,column,sheet):
    # 从PIL图像创建Image对象
    image_path = file_name # 替换为你的图片路径
    # image_path = '/'+ image_path

    img = Image(image_path)
    sheet.add_image(img,chr(ord('A')+column-1) + str(row_index))
    # 调整行高和列宽以适应图片
    sheet.row_dimensions[row_index].height = img.height
    sheet.column_dimensions[chr(ord('A')+column-1)].width = 70   # 调整列宽适当地
    return sheet
    
def save_imgs(result_dir, folder_name, results, model):
    full_dir = osp.join(result_dir, folder_name)
    mmcv.mkdir_or_exist(full_dir)
    mmcv.dump(results, osp.join(full_dir, folder_name + '.json'))

    # save imgs
    show_keys = ['pred_score', 'pred_class', 'gt_class']
    for result in results:
        result_show = dict((k, v) for k, v in result.items() if k in show_keys)
        outfile = osp.join(full_dir, osp.basename(result['filename']))
        model.show_result(result['filename'], result_show, out_file=outfile)


def main():
    args = parse_args()

    # load test results
    outputs = mmcv.load(args.result)
    assert ('pred_score' in outputs and 'pred_class' in outputs
            and 'pred_label' in outputs), \
        'No "pred_label", "pred_score" or "pred_class" in result file, ' \
        'please set "--out-items" in test.py'

    cfg = mmcv.Config.fromfile(args.config)
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)

    model = build_classifier(cfg.model)

    # build the dataloader
    dataset = build_dataset(cfg.data.test)
    filenames = list()
    for info in dataset.data_infos:
        if info['img_prefix'] is not None:
            filename = osp.join(info['img_prefix'],
                                info['img_info']['filename'])
        else:
            filename = info['img_info']['filename']
        filenames.append(filename)
    gt_labels = list(dataset.get_gt_labels())
    gt_classes = [dataset.CLASSES[x] for x in gt_labels]

    outputs['filename'] = filenames
    outputs['gt_label'] = gt_labels
    outputs['gt_class'] = gt_classes

    need_keys = [
        'filename', 'gt_label', 'gt_class', 'pred_score', 'pred_label',
        'pred_class'
    ]
    outputs = {k: v for k, v in outputs.items() if k in need_keys}
    outputs_list = list()
    for i in range(len(gt_labels)):
        output = dict()
        for k in outputs.keys():
            output[k] = outputs[k][i]
        outputs_list.append(output)

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    # 获取默认的工作表
    default_sheet = workbook.active
    # 删除默认的空白工作表
    workbook.remove(default_sheet)
    # 创建一个名为 'sheet1' 的工作表
    sheet = workbook.create_sheet('sheet')
    fail = {}  # 用字典存储类别识别错误的情况和置信度
    for output in outputs_list:
        if output['pred_label'] != output['gt_label']:
            pred_class = output['pred_class']
            # pred_label = output['pred_label']
            # gt_label = output['gt_label']
            pred_score = output['pred_score']
            file_name = output['filename']
            gt_class = output['gt_class']
            if gt_class not in fail:
                fail[gt_class] = []  # 初始化对应类别的错误情况列表

            fail[gt_class].append({'pred_class': pred_class, 'pred_score': pred_score,'file_name':file_name})

    # 写入表头
    sheet.cell(row=1, column=1).value = 'Key'
    sheet.column_dimensions[chr(ord('A')+1-1)].width = 26   # 调整列宽适当地
    
    sheet.cell(row=1, column=3).value = 'pred_label'
    sheet.column_dimensions[chr(ord('A')+3-1)].width = 26   # 调整列宽适当地
    sheet.cell(row=1, column=4).value = 'picture'
    # sheet.cell(row=1, column=5).value = 'pred_score'

    sheet.cell(row=1, column=6).value = 'pred_label'
    sheet.column_dimensions[chr(ord('A')+6-1)].width = 26   # 调整列宽适当地
    sheet.cell(row=1, column=7).value = 'picture'
    # sheet.cell(row=1, column=9).value = 'pred_score'
    
    sheet.cell(row=1, column=9).value = 'pred_label'
    sheet.column_dimensions[chr(ord('A')+9-1)].width = 26   # 调整列宽适当地
    sheet.cell(row=1, column=10).value = 'picture'
    # sheet.cell(row=1, column=13).value = 'pred_score'
    # osp.join(*args.result.split('/')[:-1],*file_name.split('/')[-2:])
    # 写入数据
    row_index = 2
    for gt_class, error_list in fail.items():
        sheet.cell(row=row_index, column=1).value = gt_class
        image_path = osp.join(dataset_root,gt_class,'2.jpg')
        sheet = Insert_image(image_path,row_index,column=2,sheet=sheet)

        if len(error_list) > 1:
            # 按照pred_score进行排序，返回索引列表
            sorted_indexes = sorted(range(len(error_list)), key=lambda i: error_list[i]['pred_score'], reverse=True)
            
            #first
            pred_class = error_list[sorted_indexes[0]]['pred_class']
            # file_name  = osp.join(*args.result.split('/')[:-1],*error_list[sorted_indexes[0]]['file_name'].split('/')[-2:])          
            Fail_File_name = "Fail_"+error_list[sorted_indexes[0]]['file_name'].split('/')[-1]
            file_name  = osp.join(*args.result.split('/')[:-1],error_list[sorted_indexes[0]]['file_name'].split('/')[-2],Fail_File_name)  
            sheet.cell(row=row_index, column=3).value = pred_class
            # sheet = Insert_image('/'+file_name,row_index,column = 4,sheet = sheet)
            sheet = Insert_image(file_name,row_index,column = 4,sheet = sheet)
            # sheet.cell(row=row_index, column=5).value = pred_score

            # #seconde
            pred_class = error_list[sorted_indexes[1]]['pred_class']
            # pred_score = error_list[sorted_indexes[1]]['pred_score']  
            # file_name  = osp.join(*args.result.split('/')[:-1],*error_list[sorted_indexes[1]]['file_name'].split('/')[-2:])          
            Fail_File_name = "Fail_"+error_list[sorted_indexes[1]]['file_name'].split('/')[-1]
            file_name  = osp.join(*args.result.split('/')[:-1],error_list[sorted_indexes[1]]['file_name'].split('/')[-2],Fail_File_name)
            sheet.cell(row=row_index, column=6).value = pred_class
            sheet = Insert_image(file_name,row_index,column = 7,sheet = sheet)
            # sheet.cell(row=row_index, column=9).value = pred_score
            
            if len(error_list) > 2:
                #third
                pred_class = error_list[sorted_indexes[2]]['pred_class']
                # pred_score = error_list[sorted_indexes[2]]['pred_score']    
                # file_name  = osp.join(*args.result.split('/')[:-1],*error_list[sorted_indexes[2]]['file_name'].split('/')[-2:])          
                Fail_File_name = "Fail_"+error_list[sorted_indexes[2]]['file_name'].split('/')[-1]
                file_name  = osp.join(*args.result.split('/')[:-1],error_list[sorted_indexes[2]]['file_name'].split('/')[-2],Fail_File_name)
                sheet.cell(row=row_index, column=9).value = pred_class
                sheet = Insert_image(file_name,row_index,column = 10,sheet = sheet)
                # sheet.cell(row=row_index, column=13).value = pred_score
            row_index += 1
        else:
            pred_class = error_list[0]['pred_class']
            # pred_score = error_list[0]['pred_score']
            # file_name  = osp.join(*args.result.split('/')[:-1],*error_list[0]['file_name'].split('/')[-2:])          
            # 写入 pred_class
            Fail_File_name = "Fail_"+error_list[0]['file_name'].split('/')[-1]
            file_name  = osp.join(*args.result.split('/')[:-1],error_list[0]['file_name'].split('/')[-2],Fail_File_name)
            sheet.cell(row=row_index, column=3).value = pred_class
            sheet = Insert_image(file_name,row_index,column = 4,sheet = sheet)
            # 写入 pred_score
            # sheet.cell(row=row_index, column=5).value = pred_score
            
            row_index += 1

    # calculate confusion matrix
    class_names = dataset.CLASSES
    confusion_matrix = pd.DataFrame(0, index=class_names, columns=class_names)
    for output in outputs_list:
        gt_cls = dataset.CLASSES[output['gt_label']]
        pred_cls = dataset.CLASSES[output['pred_label']]
        confusion_matrix.loc[gt_cls, pred_cls] += 1  
    # confusion_matrix.to_excel('confusion_matrix.xlsx',sheet_name='sheet1')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    # 创建一个名为 'sheet1' 的工作表
    sheet1 = workbook.create_sheet('sheet1')

    # 获取类别名称列表
    class_names = list(dataset.CLASSES)

    # 写入表头
    header_row = [''] + class_names
    sheet1.append(header_row)

    # 计算并写入混淆矩阵数据
    for i in range(len(class_names)):
        gt_cls = class_names[i]
        row = [gt_cls]
        for j in range(len(class_names)):
            pred_cls = class_names[j]
            count = confusion_matrix.loc[gt_cls, pred_cls]
            row.append(count)
        sheet1.append(row)



    class_precisions = []
    class_recalls = []

    # calculate precision and recall for each class
    class_names = dataset.CLASSES
    class_counts = {name: 0 for name in class_names}
    true_positives = {name: 0 for name in class_names}
    false_positives = {name: 0 for name in class_names}
    false_negatives = {name: 0 for name in class_names}
    for output in outputs_list:
        gt_cls = dataset.CLASSES[output['gt_label']]
        pred_cls = dataset.CLASSES[output['pred_label']]
        class_counts[gt_cls] += 1
 
        if gt_cls == pred_cls:
            true_positives[gt_cls] += 1
        else:
            false_positives[pred_cls] += 1
            false_negatives[gt_cls] += 1

    print('Class\tPrecision\tRecall')
    for cls_name in class_names:
        tp = true_positives[cls_name]
        fp = false_positives[cls_name]
        fn = false_negatives[cls_name]
        precision = tp / (tp + fp + 1e-6)
        recall = tp / (tp + fn + 1e-6)
        print('{}\t{:.5f}\t{:.5f}'.format(cls_name, precision, recall))
        class_precisions.append(precision)
        class_recalls.append(recall)
    
    sheet2 = workbook.create_sheet('sheet2')
        # 写入表头
    sheet2['A1'] = 'Class'
    sheet2['B1'] = 'Precision'
    sheet2['C1'] = 'Recall'
    sheet2.column_dimensions[chr(ord('A'))].width = 26   # 调整列宽适当地
    # 写入数据
    for i, cls_name in enumerate(class_names):
        row = i + 2  # 从第二行开始写入数据
        sheet2.cell(row=row, column=1, value=cls_name)
        sheet2.cell(row=row, column=2, value=class_precisions[i])
        sheet2.cell(row=row, column=3, value=class_recalls[i])


    # 计算平均召回率和平均精确率
    mean_precision = np.mean(class_precisions)
    mean_recall = np.mean(class_recalls)

    print('Average Precision: {:.5f}'.format(mean_precision))
    print('Average Recall: {:.5f}'.format(mean_recall))
    # calculate average precision and recall
    accuracy = np.diag(confusion_matrix).sum() / confusion_matrix.values.sum()
    print('Accuracy: {:.5f}'.format(accuracy))
    # 保存 Excel 文件
    workbook.save('confusion_matrix.xlsx')
if __name__ == '__main__':
    main()
